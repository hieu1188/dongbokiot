"""
dongbo_kiot — Báo cáo Lãi/Lỗ (P&L) HỢP NHẤT nhiều tài khoản KiotViet.

App desktop độc lập (KHÔNG liên quan app.py). Gộp doanh thu + giá vốn của các
tài khoản, trừ phí sàn (theo % cấu hình) -> ra lãi thực theo kênh và theo sản phẩm.

Chạy:  python dongbo_kiot.py
"""
import threading
from datetime import datetime, timedelta

import customtkinter as ctk
from tkinter import messagebox, ttk, filedialog
from tkcalendar import DateEntry

import kiot_pnl as core

ORANGE = "#fe8b16"


def fmt(n):
    try:
        return f"{float(n):,.0f}"
    except (TypeError, ValueError):
        return str(n)


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("ĐỒNG BỘ KIOT — BÁO CÁO LÃI/LỖ HỢP NHẤT")
        self.geometry("1250x820")
        ctk.set_appearance_mode("Light")

        self.cfg = core.load_config()
        self._build_header()
        self._build_toolbar()
        self._build_summary()
        self._build_tables()

        if not self.cfg.get("accounts"):
            self.after(300, lambda: messagebox.showinfo(
                "Cấu hình", "Chưa có tài khoản KiotViet.\nBấm '⚙️ Cấu hình' để nhập."))

    # --------------- UI ---------------
    def _build_header(self):
        h = ctk.CTkFrame(self, fg_color="#ffffff", corner_radius=0, border_width=1)
        h.pack(fill="x", padx=16, pady=(16, 8))
        ctk.CTkLabel(h, text="LÃI / LỖ HỢP NHẤT", font=("Arial", 26, "bold"),
                     text_color=ORANGE).pack(side="left", padx=18, pady=10)
        self.status = ctk.CTkLabel(h, text="Sẵn sàng", font=("Arial", 14),
                                   text_color="#666")
        self.status.pack(side="right", padx=20)

    def _build_toolbar(self):
        t = ctk.CTkFrame(self, fg_color="#fdfdfd", corner_radius=0, border_width=1)
        t.pack(fill="x", padx=16, pady=6)

        ctk.CTkLabel(t, text="Từ:", font=("Arial", 13)).pack(side="left", padx=(14, 2), pady=14)
        self.date_from = DateEntry(t, width=11, date_pattern="dd/mm/yyyy",
                                   font=("Arial", 12), background=ORANGE,
                                   foreground="white", borderwidth=2)
        self.date_from.pack(side="left")
        self.date_from.set_date(datetime.now() - timedelta(days=30))

        ctk.CTkLabel(t, text="Đến:", font=("Arial", 13)).pack(side="left", padx=(12, 2))
        self.date_to = DateEntry(t, width=11, date_pattern="dd/mm/yyyy",
                                 font=("Arial", 12), background=ORANGE,
                                 foreground="white", borderwidth=2)
        self.date_to.pack(side="left")
        self.date_to.set_date(datetime.now())

        ctk.CTkButton(t, text="📊 CHẠY BÁO CÁO", command=self.run_report_thread,
                      fg_color=ORANGE, font=("Arial", 13, "bold"), width=160
                      ).pack(side="left", padx=16)
        ctk.CTkButton(t, text="⚙️ Cấu hình", command=self.open_settings,
                      fg_color="#555", font=("Arial", 12), width=100
                      ).pack(side="left", padx=4)
        ctk.CTkButton(t, text="⬇ Xuất Excel", command=self.export_excel,
                      fg_color="#2e7d32", font=("Arial", 12), width=110
                      ).pack(side="right", padx=14)

    def _build_summary(self):
        self.sum_frame = ctk.CTkFrame(self, fg_color="#ffffff", corner_radius=0, border_width=1)
        self.sum_frame.pack(fill="x", padx=16, pady=6)
        self.cards = {}
        specs = [("revenue", "Doanh thu", "#1565c0"),
                 ("cogs", "Giá vốn", "#6a1b9a"),
                 ("gross", "Lãi gộp", "#00838f"),
                 ("fee", "Phí sàn", "#b26a00"),
                 ("shipping", "Ship", "#8a6d00"),
                 ("net", "LÃI THỰC", "#0a7d28")]
        for key, label, color in specs:
            c = ctk.CTkFrame(self.sum_frame, fg_color="#fafafa", corner_radius=8,
                             border_width=1)
            c.pack(side="left", expand=True, fill="both", padx=6, pady=10)
            ctk.CTkLabel(c, text=label, font=("Arial", 12), text_color="#888").pack(pady=(8, 0))
            val = ctk.CTkLabel(c, text="—", font=("Arial", 20, "bold"), text_color=color)
            val.pack(pady=(0, 8))
            self.cards[key] = val
        self.margin_lbl = ctk.CTkLabel(self.sum_frame, text="", font=("Arial", 13, "bold"))
        self.margin_lbl.pack(side="left", padx=10)

    def _make_tree(self, parent, cols, widths, anchors):
        wrap = ctk.CTkFrame(parent, fg_color="transparent")
        wrap.pack(fill="both", expand=True, padx=6, pady=6)
        sb = ttk.Scrollbar(wrap); sb.pack(side="right", fill="y")
        tree = ttk.Treeview(wrap, columns=cols, show="headings", yscrollcommand=sb.set)
        for c, w, a in zip(cols, widths, anchors):
            tree.heading(c, text=c)
            tree.column(c, width=w, anchor=a)
        tree.pack(fill="both", expand=True)
        sb.config(command=tree.yview)
        tree.tag_configure("loss", foreground="#c0271c")
        tree.tag_configure("win", foreground="#0a7d28")
        return tree

    def _build_tables(self):
        self.tabs = ctk.CTkTabview(self, fg_color="#ffffff")
        self.tabs.pack(fill="both", expand=True, padx=16, pady=(6, 16))
        tab_c = self.tabs.add("Theo kênh")
        tab_p = self.tabs.add("Theo sản phẩm")

        self.tree_chan = self._make_tree(
            tab_c,
            ("Kênh", "Số đơn", "Doanh thu", "Giá vốn", "Lãi gộp", "Phí sàn", "Lãi thực", "% Lãi"),
            (170, 70, 130, 130, 130, 120, 130, 80),
            ("w", "center", "e", "e", "e", "e", "e", "e"))
        self.tree_prod = self._make_tree(
            tab_p,
            ("Mã hàng", "Tên sản phẩm", "SL bán", "Doanh thu", "Giá vốn", "Lãi gộp", "% Lãi"),
            (120, 340, 80, 130, 130, 130, 80),
            ("w", "w", "center", "e", "e", "e", "e"))

    # --------------- CHẠY BÁO CÁO ---------------
    def run_report_thread(self):
        if not self.cfg.get("accounts"):
            messagebox.showwarning("Thiếu cấu hình", "Chưa có tài khoản. Bấm '⚙️ Cấu hình'.")
            return
        threading.Thread(target=self.run_report, daemon=True).start()

    def _set_status(self, text, color="#666"):
        self.after(0, lambda: self.status.configure(text=text, text_color=color))

    def run_report(self):
        try:
            f = datetime.combine(self.date_from.get_date(), datetime.min.time())
            t = datetime.combine(self.date_to.get_date(), datetime.min.time())
            if f > t:
                f, t = t, f
            f_date = f.strftime("%Y-%m-%dT00:00:00")
            t_date = t.strftime("%Y-%m-%dT23:59:59")

            def prog(name, n):
                self._set_status(f"[{name}] đang tải… {n}", ORANGE)

            datasets = []
            for acc in self.cfg["accounts"]:
                if not acc.get("retailer"):
                    continue
                client = core.KiotClient(acc)
                self._set_status(f"[{acc.get('name')}] lấy giá vốn…", ORANGE)
                cost_map = client.fetch_cost_map(progress=prog)
                channels = client.fetch_sale_channels()
                self._set_status(f"[{acc.get('name')}] lấy hóa đơn…", ORANGE)
                invoices = client.fetch_invoices(f_date, t_date, progress=prog)
                datasets.append({"name": acc.get("name"), "invoices": invoices,
                                 "channels": channels, "cost_map": cost_map})

            summary, by_channel, by_product = core.compute_pnl(
                datasets, self.cfg.get("fees", {}),
                shipping_per_order=self.cfg.get("shipping_per_order", 0),
                internal_customers=self.cfg.get("internal_customer_names", []))

            self._last = (summary, by_channel, by_product)
            self.after(0, lambda: self._render(summary, by_channel, by_product))
            self._set_status(
                f"Xong: {summary['orders']} đơn"
                + (f" · bỏ {summary['excluded_internal']} đơn nội bộ"
                   if summary['excluded_internal'] else ""), "#0a7d28")
        except Exception as e:  # noqa
            self._set_status(f"LỖI: {e}", "#c0271c")
            self.after(0, lambda: messagebox.showerror("Lỗi", str(e)))

    def _render(self, summary, by_channel, by_product):
        for k, lbl in self.cards.items():
            lbl.configure(text=fmt(summary.get(k, 0)))
        net = summary.get("net", 0)
        self.cards["net"].configure(text_color="#0a7d28" if net >= 0 else "#c0271c")
        self.margin_lbl.configure(
            text=f"% Lãi thực: {summary.get('margin', 0):.1f}%",
            text_color="#0a7d28" if net >= 0 else "#c0271c")

        for i in self.tree_chan.get_children():
            self.tree_chan.delete(i)
        for c in by_channel:
            tag = "win" if c["net"] >= 0 else "loss"
            self.tree_chan.insert("", "end", tags=(tag,), values=(
                c["channel"], c["orders"], fmt(c["revenue"]), fmt(c["cogs"]),
                fmt(c["gross"]), fmt(c["fee"]), fmt(c["net"]), f"{c['margin']:.1f}%"))

        for i in self.tree_prod.get_children():
            self.tree_prod.delete(i)
        for p in by_product:
            tag = "win" if p["gross"] >= 0 else "loss"
            self.tree_prod.insert("", "end", tags=(tag,), values=(
                p["code"], p["name"], fmt(p["qty"]), fmt(p["revenue"]),
                fmt(p["cogs"]), fmt(p["gross"]), f"{p['margin']:.1f}%"))

    # --------------- XUẤT EXCEL ---------------
    def export_excel(self):
        if not getattr(self, "_last", None):
            messagebox.showinfo("Chưa có dữ liệu", "Hãy chạy báo cáo trước."); return
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Thiếu thư viện", "pip install openpyxl"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")],
                                            initialfile="bao_cao_lai_lo.xlsx")
        if not path:
            return
        summary, by_channel, by_product = self._last
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Tổng hợp"
        for k in ["orders", "revenue", "cogs", "gross", "fee", "shipping", "net", "margin"]:
            ws.append([k, summary.get(k, 0)])
        wc = wb.create_sheet("Theo kênh")
        wc.append(["Kênh", "Số đơn", "Doanh thu", "Giá vốn", "Lãi gộp", "Phí sàn", "Lãi thực", "% Lãi"])
        for c in by_channel:
            wc.append([c["channel"], c["orders"], c["revenue"], c["cogs"],
                       c["gross"], c["fee"], c["net"], round(c["margin"], 1)])
        wp = wb.create_sheet("Theo sản phẩm")
        wp.append(["Mã hàng", "Tên", "SL bán", "Doanh thu", "Giá vốn", "Lãi gộp", "% Lãi"])
        for p in by_product:
            wp.append([p["code"], p["name"], p["qty"], p["revenue"],
                       p["cogs"], p["gross"], round(p["margin"], 1)])
        wb.save(path)
        messagebox.showinfo("Đã xuất", f"Đã lưu:\n{path}")

    # --------------- CẤU HÌNH ---------------
    def open_settings(self):
        win = ctk.CTkToplevel(self)
        win.title("Cấu hình")
        win.geometry("560x640")
        win.transient(self); win.grab_set()

        scroll = ctk.CTkScrollableFrame(win)
        scroll.pack(fill="both", expand=True, padx=16, pady=12)

        accounts = self.cfg.get("accounts", [])
        while len(accounts) < 2:
            accounts.append({"name": f"TK{len(accounts)+1}", "retailer": "",
                             "client_id": "", "client_secret": ""})
        acc_entries = []
        for i, acc in enumerate(accounts):
            ctk.CTkLabel(scroll, text=f"— Tài khoản {i+1} —",
                         font=("Arial", 14, "bold"), text_color=ORANGE).pack(anchor="w", pady=(10, 2))
            e = {}
            for field, label in [("name", "Tên gợi nhớ"), ("retailer", "Retailer"),
                                 ("client_id", "Client ID"), ("client_secret", "Client Secret")]:
                ctk.CTkLabel(scroll, text=label, font=("Arial", 12)).pack(anchor="w")
                ent = ctk.CTkEntry(scroll, width=500)
                ent.insert(0, acc.get(field, ""))
                ent.pack(anchor="w", pady=(0, 4))
                e[field] = ent
            acc_entries.append(e)

        fees = self.cfg.get("fees", {"Shopee": 23, "TikTok": 23, "default": 0})
        ctk.CTkLabel(scroll, text="— Phí sàn (%) —", font=("Arial", 14, "bold"),
                     text_color=ORANGE).pack(anchor="w", pady=(12, 2))
        fee_entries = {}
        for key in ["Shopee", "TikTok", "default"]:
            ctk.CTkLabel(scroll, text=f"{key}", font=("Arial", 12)).pack(anchor="w")
            ent = ctk.CTkEntry(scroll, width=120)
            ent.insert(0, str(fees.get(key, 0)))
            ent.pack(anchor="w", pady=(0, 4))
            fee_entries[key] = ent

        ctk.CTkLabel(scroll, text="Phí ship mỗi đơn (bạn chịu, VND)",
                     font=("Arial", 12)).pack(anchor="w", pady=(8, 0))
        ship_ent = ctk.CTkEntry(scroll, width=160)
        ship_ent.insert(0, str(self.cfg.get("shipping_per_order", 0)))
        ship_ent.pack(anchor="w", pady=(0, 4))

        ctk.CTkLabel(scroll, text="Tên khách 'nội bộ' cần loại (cách nhau bởi dấu phẩy)",
                     font=("Arial", 12)).pack(anchor="w", pady=(8, 0))
        internal_ent = ctk.CTkEntry(scroll, width=500)
        internal_ent.insert(0, ", ".join(self.cfg.get("internal_customer_names", [])))
        internal_ent.pack(anchor="w", pady=(0, 4))

        def save():
            new_accounts = []
            for e in acc_entries:
                vals = {k: e[k].get().strip() for k in e}
                if vals["retailer"]:
                    new_accounts.append(vals)
            new_fees = {}
            for k, ent in fee_entries.items():
                try:
                    new_fees[k] = float(ent.get().strip() or 0)
                except ValueError:
                    new_fees[k] = 0
            try:
                ship = float(ship_ent.get().strip() or 0)
            except ValueError:
                ship = 0
            internal = [s.strip() for s in internal_ent.get().split(",") if s.strip()]
            self.cfg = {"accounts": new_accounts, "fees": new_fees,
                        "shipping_per_order": ship, "internal_customer_names": internal}
            core.save_config(self.cfg)
            win.destroy()
            messagebox.showinfo("Đã lưu", "Đã lưu cấu hình.")

        ctk.CTkButton(win, text="💾 LƯU", command=save, fg_color=ORANGE,
                      font=("Arial", 14, "bold"), height=42).pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkButton(win, text="Đóng", command=win.destroy, fg_color="#888",
                      height=34).pack(fill="x", padx=16, pady=(0, 14))


if __name__ == "__main__":
    App().mainloop()
